import os
import re
from datetime import datetime
from openpyxl import Workbook
import pdfplumber

# MyBot.py

class MeuBot:
    def __init__(self):
        pass

    def processar_faturas(self):
        # Configurações iniciais
        diretorio_pdf = 'pdf_faturas'
        pasta_saida = 'Listagem_Faturas'

        # Verificar existência de arquivos PDF
        arquivos = os.listdir(diretorio_pdf)
        if not arquivos:
            raise FileNotFoundError("Nenhum arquivo foi encontrado no diretório especificado.")

        # Criar planilha
        wb = Workbook()
        ws = wb.active
        ws.title = 'Importação de Faturas'

        # Cabeçalho
        cabecalhos = [
            'Fatura', 'Data de Emissão', 'Data de Vencimento', 'Valor da Fatura',
            'Nome do Fornecedor', 'CNPJ do Fornecedor', 'Endereço do Fornecedor', 'Arquivo da Fatura'
        ]
        for i, titulo in enumerate(cabecalhos, start=1):
            ws.cell(row=1, column=i, value=titulo)

        # Começa da próxima linha após o cabeçalho
        linha_excel = 2

        # Regex patterns
        padrao_fatura = r'FATURA # (\d+)'
        padrao_emissao = r'DATA DE EMISSÃO (\d{2}/\d{2}/\d{4})'
        padrao_vencimento = r'DATA DE VENCIMENTO (\d{2}/\d{2}/\d{4})'

        # Processar cada PDF
        for arquivo in arquivos:
            caminho_pdf = os.path.join(diretorio_pdf, arquivo)
            with pdfplumber.open(caminho_pdf) as pdf:
                texto = pdf.pages[0].extract_text()
            
            linhas = texto.splitlines()

            # Inicialização dos dados
            fatura = 'Não encontrado'
            emissao = 'Não encontrado'
            vencimento = 'Não encontrado'
            valor = 'Não encontrado'
            fornecedor = 'Não encontrado'
            cnpj = 'Não encontrado'
            endereco = 'Não encontrado'

            for i, linha in enumerate(linhas):
                # print(f'{i} - {linha}')
                if re.search(padrao_fatura, linha):
                    fatura = re.search(padrao_fatura, linha).group(1)
                elif re.search(padrao_emissao, linha):
                    emissao = re.search(padrao_emissao, linha).group(1)
                elif re.search(padrao_vencimento, linha):
                    vencimento = re.search(padrao_vencimento, linha).group(1)
                elif 'R$' in linha:
                    valor_split = linha.split('R$')
                    if len(valor_split) > 1:
                        valor = valor_split[1].strip()

            # Tenta capturar os dados do fornecedor (linhas 0 a 3)
            try:
                fornecedor = linhas[0]
                cnpj = linhas[1]
                endereco = f"{linhas[2]}, {linhas[3]}"
            except IndexError:
                pass  # Deixa como "Não encontrado" se faltar informação

            # Preencher planilha
            dados = [fatura, emissao, vencimento, valor, fornecedor, cnpj, endereco, arquivo]
            for col, dado in enumerate(dados, start=1):
                ws.cell(row=linha_excel, column=col, value=dado)

            linha_excel += 1

        # Gerar nome do arquivo com data/hora
        data_hora = datetime.now().strftime('%Y-%m-%d %H-%M-%S')
        os.makedirs(pasta_saida, exist_ok=True)
        caminho_saida = os.path.join(pasta_saida, f'Faturas - {data_hora}.xlsx')
        wb.save(caminho_saida)

        print(f"Planilha gerada com sucesso: {caminho_saida}")
