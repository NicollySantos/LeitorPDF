#importação de bibliotecas
import os
#navegação pelo diretório
from openpyxl import Workbook
#excel
import pdfplumber
#leitura e interação com o pdf
from datetime import datetime 
#importação de data e hora
import re
#instrução regex

directory = 'pdf_faturas'
files = os.listdir(directory)
files_quant = len(files)

#idenfica a quantidade de arquivos

if files_quant == 0:
    raise Exception("Não foram encontrados arquivos no diretório")
#se for igual a 0, sem arquivos encontrados no diretório

#criando a planilha
wb = Workbook()
#sheet
ws = wb.active
#título da sheet
ws.title = 'Importação de Faturas'


ws['A1'] = 'Fatura'
ws['B1'] = 'Data de Emissão'
ws['C1'] = 'Data de Vencimento'
ws['D1'] = 'Valor da Fatura'
ws['E1'] = 'Nome do Fornecedor'
ws['F1'] = 'Slogan do Fornecedor'
ws['G1'] = 'Arquivo da Fatura'

Last_empty_line = 1
while ws['A' + str(Last_empty_line)].value is not None:
    Last_empty_line += 1

for file in files:
    with pdfplumber.open(directory + "/" + file) as pdf:
    #abrindo os pdfs encontrados na pasta
        first_page = pdf.pages[0]
        pdf_text = first_page.extract_text()
        # print(pdf_text)

    #Instrução Regex permite compreender expressões regulares, manipulando textos com bases em padrões específicos
    inv_numero_re_patterns = r'FATURA # (\d+)'
    inv_dataE_re_pattern = r'DATA DE EMISSÃO (\d{2}/\d{2}/\d{4})'
    inv_dataV_re_pattern = r'DATA DE VENCIMENTO (\d{2}/\d{2}/\d{4})'

    linhas = pdf_text.splitlines()
    invoice_numero = 'Não encontrado'
    invoice_dataE = 'Não encontrado'
    invoice_dataV = 'Não encontrado'
    invoice_valorTotal = 'Não encontrado'
    invoice_valorNomeFornecedor = 'Não encontrado'
    invoice_valorSloganFornecedor = 'Não encontrado'
    
    for index, linha in enumerate(linhas):
        #print(f'{index} - {linha}')
        if(re.search(inv_numero_re_patterns, linha)):
            invoice_numero =re.search(inv_numero_re_patterns, linha).group(1) 
        elif(re.search('FATURA', linha)):
            invoice_numero = linhas[index-1]
        elif(re.search(inv_dataE_re_pattern, linha)):
            invoice_dataE = re.search(inv_dataE_re_pattern, linha).group(1)
        elif(re.search(inv_dataV_re_pattern, linha)):
            invoice_dataV = re.search(inv_dataV_re_pattern, linha).group(1)
        elif(linha.find('$')!= -1):
            invoice_valorTotal = linha.split('R$')[1]
        
    ws['A{}'.format(Last_empty_line)] = invoice_numero
    ws['B{}'.format(Last_empty_line)] = invoice_dataE
    ws['C{}'.format(Last_empty_line)] = invoice_dataV
    ws['D{}'.format(Last_empty_line)] = invoice_valorTotal
    ws['E{}'.format(Last_empty_line)] = linhas[0]
    ws['F{}'.format(Last_empty_line)] = linhas[1]        
    ws['G{}'.format(Last_empty_line)] = file

    Last_empty_line += 1

full_now = str(datetime.now()).replace(":","-")
dot_index = full_now.index(".")
now = full_now[:dot_index]

pasta_faturas = 'Listagem_Faturas'
os.makedirs(pasta_faturas, exist_ok=True)
wb.save('{}/Faturas - {}.xlsx'.format(pasta_faturas,now))

